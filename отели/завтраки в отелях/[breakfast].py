from openpyxl import *
import os
from pathlib import Path
import zipfile

text='''Для корректной работы программы:
Сохраняйте файлы в одну папку с программой
Не кладите в папку более одного файла одного и того же отеля
Не прогоняйте программу по одному и тому же файлу повторно
Не переименовывайте исходные названия эксель файлов
Сохраняйте в папке оба файла Суперотеля, а не по отдельности\n'''
print(text)

for file in os.scandir():
    if Path(file).suffix==".zip":

        zipDirPath = file.name

        with zipfile.ZipFile(zipDirPath) as zf:
            zf.extractall()

        os.remove(file)


for file in os.scandir():
    if Path(file).suffix == ".xlsx":

        if file.name.startswith("29118"):
            try:
                os.rename(file, "Изумруд.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")


                
        elif file.name.startswith("29120"):
            try:
                os.rename(file, "Сапфир.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

         
        elif file.name.startswith("29122"):
            try:
                os.rename(file, "Рубин.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29125"):
            try:
                os.rename(file, "Бриллиант.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29127"):
            try:
                os.rename(file, "Алмаз.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29129"):
            try:
                os.rename(file, "Янтарь.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29131"):
            try:
                os.rename(file, "Жемчужина.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29123"):
            try:
                os.rename(file, "Русалка.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29124"):
            try:
                os.rename(file, "Суперотель 2 этаж.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")

                      
        elif file.name.startswith("29126"):
            try:
                os.rename(file, "Суперотель 1 этаж.xlsx")
            except FileExistsError:
                print("Файл " + file.name + " уже существует! Пропуск файла.")



for file in os.scandir():
    if Path(file).suffix == ".xlsx":
        try:
            workbook = load_workbook(file)
        except FileNotFoundError:
            print("Файл " + file.name + " не найден.")
            continue
        except Exception as e:
            print("ошибка при открытии файла " + file.name + " {e}")
            continue
        sheet = workbook.active

        sheet.delete_cols(1, 7)
        sheet.delete_cols(2, 1)
        sheet.delete_cols(3, 5)
        sheet.delete_cols(4, 2)

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 30

        
        columnRooms = sheet['B']
        for item in columnRooms:
   
            item.value=str(item.value).replace("№", "")
            item.value=str(item.value).replace(" РАЗД", "")
            item.value=str(item.value).replace(" ДВСП", "")
            item.value=str(item.value).replace(" ОДНОСП", "")
            item.value=str(item.value).replace(" разд", "")
            item.value=str(item.value).replace(" двсп", "")
            item.value=str(item.value).replace(" ДВУСП", "")
            item.value=str(item.value).replace(" ДВУСП+ОДНОСП", "")
            item.value=str(item.value).replace(" (ДВА ОКНА)", "")
            item.value=str(item.value).replace(" ДВУСП+ОДНОСП", "")
            item.value=str(item.value).replace("+", "")
            item.value=str(item.value).replace("ОДНОСП", "")
            



        sheet.move_range("A1:A150", rows=0, cols=3)
        sheet.move_range("B1:b150", rows=0, cols=-1)
        sheet.move_range("D1:D150", rows=0, cols=-2)

        workbook.save(file)
        workbook.close()


        try:
            workbook = load_workbook(file)
        except FileNotFoundError:
            print("Файл " + file.name + " не найден.")
            continue
        except Exception as e:
            print("ошибка при открытии файла " + file.name + " {e}")
            continue
        sheet = workbook.active
        data = []

            
        
        sheet.delete_rows(1)

        for row in sheet.values:
            for value in row:
                data.append(value)


        sheet.delete_rows(1, sheet.max_row)


        data = [data[i:i+3] for i in range(0, len(data), 3)]

        
        for i in range(len(data)):
            data[i][0] = int(data[i][0])
        data = sorted(data, key = lambda x: x[0])


        sheet.append(("Номер", "Тариф", "Кол-во человек"))
        for row in data:
            sheet.append(row)
    
      
        workbook.save(file)
        workbook.close()

        outputfilename=str(file).replace('<DirEntry ', '')
        outputfilename=outputfilename.replace('>', '')
        print('Файл ', outputfilename, ' создан.')


if ((os.path.exists("Суперотель 2 этаж.xlsx")) & (os.path.exists("Суперотель 1 этаж.xlsx"))):
    try:
        workbook = load_workbook("Суперотель 2 этаж.xlsx")
    except FileNotFoundError:
        print("Файл Суперотель 2 этаж" + " не найден.")

    except Exception as e:
        print("ошибка при открытии файла Суперотель 2 этаж" + " {e}")

 
    sheet = workbook.active
    sheet.delete_rows(1)
    n95_data=[]

    for row in sheet.values:
        for value in row:
            n95_data.append(value)

    workbook.close()


    try:
        workbook = load_workbook("Суперотель 1 этаж.xlsx")
    except FileNotFoundError:
        print("Файл Суперотель 1 этаж" +  " не найден.")

    except Exception as e:
        print("ошибка при открытии файла Суперотель 1 этаж" + " {e}")

    sheet = workbook.active
    sheet.delete_rows(1)

    for row in sheet.values:
        for value in row:
            n95_data.append(value)


    sheet.delete_rows(1, sheet.max_row)


    n95_data = [n95_data[i:i+3] for i in range(0, len(n95_data), 3)]
    for i in range(len(n95_data)):
        n95_data[i][0] = int(n95_data[i][0])
    n95_data = sorted(n95_data, key = lambda x: x[0])


    sheet.append(("Номер", "Тариф", "Кол-во человек"))
    for row in n95_data:
        sheet.append(row)
    
    workbook.save("Суперотель.xlsx")
    workbook.close()



    try:
        os.remove("Суперотель 1 этаж.xlsx")
        os.remove("Суперотель 2 этаж.xlsx")
        print('\nФайлы Суперотеля успешно объединены')
    except FileNotFoundError:
        print("Не удалось удалить файлы Суперотеля, так как хотя бы один из них не существует!")

print('\nКонец работы программы\n\nЧтобы закрыть окно нажмите Enter')
input()



