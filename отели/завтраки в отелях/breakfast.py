from openpyxl import *
import os
from pathlib import Path
import zipfile


dirpath=r"C:\Py"

#Распаковываем все эксель файлы из зип файлов
for file in os.scandir(dirpath):
    if Path(file).suffix==".zip":
        zipDirPath = dirpath + "\\" + file.name
        with zipfile.ZipFile(zipDirPath) as zf:
            zf.extractall(dirpath)
        os.remove(file)

#определяем отель, файл которого открыт в данный момент и
#присваем ему сответствующее название
for file in os.scandir(dirpath):
    if Path(file).suffix == ".xlsx":

        newFilename=file.name

        if file.name.startswith("29118"):
            newFileName="Изумруд.xlsx"

        elif file.name.startswith("29120"):
            newFileName="Рубин.xlsx"

        elif file.name.startswith("29122"):
                newFileName="Сапфир.xlsx"

        elif file.name.startswith("29125"):
                newFileName="Бриллиант.xlsx"
      
        elif file.name.startswith("29127"):
                newFileName="Алмаз.xlsx"
   
        elif file.name.startswith("29129"):
                 newFileName="Аквамарин.xlsx"

        elif file.name.startswith("29131"):
                newFileName="Дипломат.xlsx"
     
        elif file.name.startswith("29123"):
                newFileName="Фантастика.xlsx"
    
        elif file.name.startswith("29124"):
                newFileName="СуперОтель 1этаж.xlsx"
     
        elif file.name.startswith("29126"):
            newFileName="СуперОтель 2этаж.xlsx"

        try:
            os.rename(file, newFileName)
        except FileExistsError:
            print("Файл " + newFileName.name + " уже существует! Пропуск файла.")

#открываем каждый эксель файл и обрабатываем его
for file in os.scandir(dirpath):
    if Path(file).suffix == ".xlsx":
        try:
            workbook = load_workbook(file)
        except FileNotFoundError:
            print("Файл " + file.name + " не найден.")
            continue
        except Exception as e:
            print("ошибка при открытии файла " + file.name + " {e}")
            continue
        sheet = workbook.active

#удаляем ненужные колонки
        sheet.delete_cols(1, 7)
        sheet.delete_cols(2, 1)
        sheet.delete_cols(3, 5)
        sheet.delete_cols(4, 2)

#задаем ширину нужных колонок
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 30

#в колонке с номерами комнат содержаться дополнительные данные о типе кроватей
#приводим данные к числу, удаляя строковые символы
        columnRooms = sheet['B']
        for item in columnRooms:
            item.value=str(item.value).replace("№", "")
            item.value=str(item.value).replace(" РАЗД", "")
            item.value=str(item.value).replace(" ДВСП", "")
            item.value=str(item.value).replace(" ОДНОСП", "")
            item.value=str(item.value).replace(" разд", "")
            item.value=str(item.value).replace(" двсп", "")
            item.value=str(item.value).replace(" ДВУСП", "")
            item.value=str(item.value).replace(" ДВУСП+ОДНОСП", "")
            item.value=str(item.value).replace(" (ДВА ОКНА)", "")
            item.value=str(item.value).replace(" ДВУСП+ОДНОСП", "")
            item.value=str(item.value).replace("+", "")
            item.value=str(item.value).replace("ОДНОСП", "")
            
#меняем колонки местами
        sheet.move_range("A1:A150", rows=0, cols=3)
        sheet.move_range("B1:b150", rows=0, cols=-1)
        sheet.move_range("D1:D150", rows=0, cols=-2)

#сохраняем файл и закрываем его, потом открываем его снова
#это нужно, чтобы убрать NaN значения из памяти
#без NaN значений нам удобнее работать с данными и сортировать числа
        workbook.save(file)
        workbook.close()
        try:
            workbook = load_workbook(file)
        except FileNotFoundError:
            print("Файл " + file.name + " не найден.")
            continue
        except Exception as e:
            print("ошибка при открытии файла " + file.name + " {e}")
            continue
        sheet = workbook.active

#сортируем таблицу по возрастанию номера комнаты и сохраяем файл
        data = []
        sheet.delete_rows(1)
        for row in sheet.values:
            for value in row:
                data.append(value)

        sheet.delete_rows(1, sheet.max_row)

        data = [data[i:i+3] for i in range(0, len(data), 3)]

        for i in range(len(data)):
            data[i][0] = int(data[i][0])
        data = sorted(data, key = lambda x: x[0])

        sheet.append(("Номер", "Тариф", "Кол-во человек"))
        for row in data:
            sheet.append(row)
    
        workbook.save(file)
        workbook.close()

#Файлы СуперОтель 1-ый и 2-й этаж необходимо объединить в единый файл:
if ((os.path.exists("C:\\Py\\СуперОтель 1этаж.xlsx")) & (os.path.exists("C:\\Py\\СуперОтель 2этаж.xlsx"))):
    try:
        workbook = load_workbook("C:\\Py\\СуперОтель 2этаж.xlsx")
    except FileNotFoundError:
        print("Файл второго этажа не найден.")
    except Exception as e:
        print("ошибка при открытии файла второго этажа" + " {e}")

    sheet = workbook.active
    sheet.delete_rows(1)
    n95_data=[]

    for row in sheet.values:
        for value in row:
            n95_data.append(value)
    workbook.close()

    try:
        workbook = load_workbook("C:\\Py\\СуперОтель 1этаж.xlsx")
    except FileNotFoundError:
        print("Файл первого этажа не найден.")
    except Exception as e:
        print("ошибка при открытии файла" + " {e}")

    sheet = workbook.active
    sheet.delete_rows(1)

    for row in sheet.values:
        for value in row:
            n95_data.append(value)

    sheet.delete_rows(1, sheet.max_row)

    n95_data = [n95_data[i:i+3] for i in range(0, len(n95_data), 3)]
    for i in range(len(n95_data)):
        n95_data[i][0] = int(n95_data[i][0])
    n95_data = sorted(n95_data, key = lambda x: x[0])

    sheet.append(("Номер", "Тариф", "Кол-во человек"))
    for row in n95_data:
        sheet.append(row)
    
    workbook.save("C:\\Py\\СуперОтель.xlsx")
    workbook.close()

    try:
        os.remove("C:\\Py\\СуперОтель 1этаж.xlsx")
        os.remove("C:\\Py\\СуперОтель 2этаж.xlsx")
    except FileNotFoundError:
        print("Не удалось удалить изначальные файлы СуперОтеля, так как хотя бы один из них не существует!")



